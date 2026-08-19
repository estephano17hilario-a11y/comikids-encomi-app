import json
import base64
import os
import re
import openpyxl

base_path = 'public/Formato-Pro-Masivo-2026_08_15_02.xlsx'
if os.path.exists(base_path):
    wb = openpyxl.load_workbook(base_path)
    print("Sheets in public file:", wb.sheetnames)
    print("Size of base file:", os.path.getsize(base_path), "bytes")
    
    ws1 = wb['Hoja1']
    print("Hoja1 max_row:", ws1.max_row, "max_column:", ws1.max_column)
    for r in range(1, 10):
        row_vals = [ws1.cell(r, c).value for c in range(1, 14)]
        print(f"Row {r}:", row_vals)

with open('src/data/shalom_template_base64.ts', 'r', encoding='utf-8') as f:
    b64_file_content = f.read()

match = re.search(r'SHALOM_OFFICIAL_TEMPLATE_BASE64\s*=\s*[\'"`]([^\'"`]+)[\'"`]', b64_file_content)
if match:
    b64_data = match.group(1)
    decoded = base64.b64decode(b64_data)
    print("Decoded base64 template size:", len(decoded), "bytes")
    with open('scripts/test_decoded_template.xlsx', 'wb') as out_f:
        out_f.write(decoded)
    wb_b64 = openpyxl.load_workbook('scripts/test_decoded_template.xlsx')
    print("Sheets in base64 template:", wb_b64.sheetnames)
    ws_b64 = wb_b64['Hoja1']
    print("Hoja1 in base64 max_row:", ws_b64.max_row)
